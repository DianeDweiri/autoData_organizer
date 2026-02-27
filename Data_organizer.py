from pathlib import Path
import pandas as pd
import shutil
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
import json
import sys


CONFIG_PATH = Path(__file__).parent / "config.json"

with open(CONFIG_PATH) as f:
    config = json.load(f)

BASE = Path.home() / config["base_folder"]
INPUT = BASE / config["input_folder"]
CLEANED = BASE / config["cleaned_folder"]
REPORTS = BASE / config["reports_folder"]
LOGS = BASE / config["logs_folder"]

green_threshold = config["thresholds"]["green"]
red_threshold = config["thresholds"]["red"]
columns_to_keep = config["columns_to_keep"]

for folder in [INPUT, CLEANED, REPORTS, LOGS]:
    folder.mkdir(parents=True, exist_ok=True)


logging.basicConfig(
    filename=LOGS / "automation.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logging.info(" Automation System Started:")


excel_files = list(INPUT.glob("*.xlsx"))

if not excel_files:
    logging.warning("No Excel files found in input folder.")
    print("⚠ No Excel files found in input folder.")
    sys.exit()


for file in excel_files:
    try:
        logging.info(f"Processing file: {file.name}")
        df = pd.read_excel(file)


        df = df.dropna()
        df = df.drop_duplicates()


        df = df[[col for col in columns_to_keep if col in df.columns]]


        cleaned_file = CLEANED / file.name
        df.to_excel(cleaned_file, index=False)
        logging.info(f"Cleaned file saved: {cleaned_file.name}")


        by_region = df.groupby("Region")["TotalPrice"].sum()
        by_product = df.groupby("Product")["TotalPrice"].sum()

        report_file = REPORTS / f"REPORT_{file.name}"
        with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
            by_region.to_excel(writer, sheet_name="Sales_By_Region")
            by_product.to_excel(writer, sheet_name="Sales_By_Product")

        logging.info(f"Report created: {report_file.name}")


        wb = load_workbook(report_file)
        for ws in wb.worksheets:

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align
                cell.border = border


            for column in ws.columns:
                max_len = 0
                col = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col].width = max_len + 3


            green_rule = CellIsRule(
                operator='greaterThan',
                formula=[str(green_threshold)],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            )
            red_rule = CellIsRule(
                operator='lessThan',
                formula=[str(red_threshold)],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            )

            ws.conditional_formatting.add(f"B2:B{ws.max_row}", green_rule)
            ws.conditional_formatting.add(f"B2:B{ws.max_row}", red_rule)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        wb.save(report_file)
        logging.info("Formatting applied successfully")


        shutil.move(file, INPUT / f"PROCESSED_{file.name}")
        logging.info(f"Original file archived: {file.name}")

    except Exception as e:
        logging.error(f"Error processing {file.name} -> {e}")

logging.info("Automation System Finished Successfully")

print(" FULL DATA AUTOMATION PIPELINE COMPLETED")